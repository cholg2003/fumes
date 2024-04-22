import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

class StudentAccountingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Accounting System")
        
        self.labels = ["Admission Number", "Name", "Class", "Term", "Total Fees", "Fees Paid"]
        self.entries = {}
        
        # Create labels and entry fields for student information
        for i, label_text in enumerate(self.labels):
            label = tk.Label(root, text=label_text)
            label.grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entry = tk.Entry(root)
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries[label_text] = entry
        
        # Dropdown for selecting term
        self.term_options = ["1st Term", "2nd Term", "3rd Term"]
        self.term_var = tk.StringVar(root)
        self.term_var.set(self.term_options[0])  # Default value
        self.term_dropdown = tk.OptionMenu(root, self.term_var, *self.term_options)
        self.term_dropdown.grid(row=self.labels.index("Term"), column=1, padx=5, pady=5)
        
        # Button for adding student
        self.add_button = tk.Button(root, text="Add Student", command=self.add_student)
        self.add_button.grid(row=len(self.labels), column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        # Button for generating Excel
        self.generate_excel_button = tk.Button(root, text="Generate Excel", command=self.generate_excel)
        self.generate_excel_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        # Initialize workbook for Excel file
        self.wb = Workbook()
        self.class_sheets = {}  # Dictionary to store sheets for each class
    
    def clear_entries(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
    
    def add_student(self):
        # Function to add a student record
        admission_number = self.entries["Admission Number"].get()
        name = self.entries["Name"].get()
        class_ = self.entries["Class"].get()
        term = self.term_var.get()
        total_fees = self.entries["Total Fees"].get()
        fees_paid = self.entries["Fees Paid"].get()
        balance = int(total_fees) - int(fees_paid)
        
        # Create a new sheet for the class if it doesn't exist
        if class_ not in self.class_sheets:
            self.class_sheets[class_] = self.wb.create_sheet(class_)
            self.class_sheets[class_].append(["Admission Number", "Name", "Term", "Total Fees", "Fees Paid", "Balance"])
        
        self.class_sheets[class_].append([admission_number, name, term, total_fees, fees_paid, balance])
        messagebox.showinfo("Success", "Student record added successfully!")
        self.clear_entries()
    
    def generate_excel(self):
        # Function to generate an Excel file with student data
        self.wb.remove(self.wb.active)  # Remove the default sheet
        self.wb.save("student_data.xlsx")
        messagebox.showinfo("Success", "Excel file generated and saved successfully!")

if __name__ == "__main__":
    root = tk.Tk()
    app = StudentAccountingApp(root)
    root.mainloop()
