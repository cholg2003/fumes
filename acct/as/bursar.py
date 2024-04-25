import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

class StudentAccountingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Accounting System")
        
        self.labels = ["Name", "Class", "Total Fees", "Fees Paid"]
        self.entries = {}
        
        # Create labels and entry fields for student information
        for i, label_text in enumerate(self.labels):
            label = tk.Label(root, text=label_text)
            label.grid(row=i, column=0, padx=5, pady=5, sticky="w")
            entry = tk.Entry(root)
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries[label_text] = entry
        
        # Buttons for operations
        self.add_button = tk.Button(root, text="Add Student", command=self.add_student)
        self.add_button.grid(row=len(self.labels), column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        self.search_button = tk.Button(root, text="Search Student", command=self.search_student)
        self.search_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        self.update_button = tk.Button(root, text="Update Student", command=self.update_student)
        self.update_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        self.delete_button = tk.Button(root, text="Delete Student", command=self.delete_student)
        self.delete_button.grid(row=len(self.labels) + 3, column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        self.generate_excel_button = tk.Button(root, text="Generate Excel", command=self.generate_excel)
        self.generate_excel_button.grid(row=len(self.labels) + 4, column=0, columnspan=2, padx=5, pady=5, sticky="we")
        
        # Initialize workbook for Excel file
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(["Name", "Class", "Total Fees", "Fees Paid", "Balance"])
    
    def clear_entries(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
    
    def add_student(self):
        # Function to add a student record
        name = self.entries["Name"].get()
        class_ = self.entries["Class"].get()
        total_fees = self.entries["Total Fees"].get()
        fees_paid = self.entries["Fees Paid"].get()
        balance = int(total_fees) - int(fees_paid)
        
        self.ws.append([name, class_, total_fees, fees_paid, balance])
        messagebox.showinfo("Success", "Student record added successfully!")
        self.clear_entries()
    
    def search_student(self):
        # Function to search for a student record
        name_to_search = self.entries["Name"].get()
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=1, max_col=1):
            if row[0].value == name_to_search:
                messagebox.showinfo("Student Found", f"Name: {row[0].value}, Class: {row[1].value}, Total Fees: {row[2].value}, Fees Paid: {row[3].value}, Balance: {row[4].value}")
                return
        messagebox.showinfo("Student Not Found", "Student not found in records.")
    
    def update_student(self):
        # Function to update a student record
        name_to_update = self.entries["Name"].get()
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=1, max_col=1):
            if row[0].value == name_to_update:
                class_ = self.entries["Class"].get()
                total_fees = self.entries["Total Fees"].get()
                fees_paid = self.entries["Fees Paid"].get()
                balance = int(total_fees) - int(fees_paid)
                
                row[1].value = class_
                row[2].value = total_fees
                row[3].value = fees_paid
                row[4].value = balance
                
                messagebox.showinfo("Success", "Student record updated successfully!")
                self.clear_entries()
                return
        messagebox.showinfo("Student Not Found", "Student not found in records.")
    
    def delete_student(self):
        # Function to delete a student record
        name_to_delete = self.entries["Name"].get()
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=1, max_col=1):
            if row[0].value == name_to_delete:
                self.ws.delete_rows(row[0].row)
                messagebox.showinfo("Success", "Student record deleted successfully!")
                self.clear_entries()
                return
        messagebox.showinfo("Student Not Found", "Student not found in records.")
    
    def generate_excel(self):
        # Function to generate an Excel file with student data
        self.wb.save("student_data.xlsx")
        messagebox.showinfo("Success", "Excel file generated and saved successfully!")

if __name__ == "__main__":
    root = tk.Tk()
    app = StudentAccountingApp(root)
    root.mainloop()
