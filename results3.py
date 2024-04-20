import tkinter as tk
from tkinter import messagebox, simpledialog
import sqlite3
from openpyxl import load_workbook

class ExamResultsSystem:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Primary Pupil Exam Results System")

        self.conn = sqlite3.connect('exam_results.db')
        self.create_table()

        self.create_widgets()

    def create_table(self):
        c = self.conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS results
                     (name TEXT PRIMARY KEY, english INTEGER, mathematics INTEGER,
                     science INTEGER, cre INTEGER, sst INTEGER)''')
        self.conn.commit()

    def create_widgets(self):
        # Input Section
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=10)

        tk.Label(input_frame, text="Name:").grid(row=0, column=0)
        self.name_entry = tk.Entry(input_frame)
        self.name_entry.grid(row=0, column=1)

        tk.Label(input_frame, text="English:").grid(row=1, column=0)
        self.english_entry = tk.Entry(input_frame)
        self.english_entry.grid(row=1, column=1)

        tk.Label(input_frame, text="Mathematics:").grid(row=2, column=0)
        self.math_entry = tk.Entry(input_frame)
        self.math_entry.grid(row=2, column=1)

        tk.Label(input_frame, text="Science:").grid(row=3, column=0)
        self.science_entry = tk.Entry(input_frame)
        self.science_entry.grid(row=3, column=1)

        tk.Label(input_frame, text="C.R.E:").grid(row=4, column=0)
        self.cre_entry = tk.Entry(input_frame)
        self.cre_entry.grid(row=4, column=1)

        tk.Label(input_frame, text="S.S.T:").grid(row=5, column=0)
        self.sst_entry = tk.Entry(input_frame)
        self.sst_entry.grid(row=5, column=1)

        add_btn = tk.Button(input_frame, text="Add Result", command=self.add_result)
        add_btn.grid(row=6, columnspan=2, pady=5)

        # Report Section
        report_frame = tk.Frame(self.root)
        report_frame.pack()

        self.report_text = tk.Text(report_frame, width=60, height=10)
        self.report_text.pack()

        generate_report_btn = tk.Button(report_frame, text="Generate Report", command=self.generate_report)
        generate_report_btn.pack(pady=5)

    def add_result(self):
        name = self.name_entry.get()
        english = self.validate_marks(self.english_entry.get())
        math = self.validate_marks(self.math_entry.get())
        science = self.validate_marks(self.science_entry.get())
        cre = self.validate_marks(self.cre_entry.get())
        sst = self.validate_marks(self.sst_entry.get())

        if not name or english is None or math is None or science is None or cre is None or sst is None:
            messagebox.showerror("Error", "Please enter valid data for all fields.")
            return

        c = self.conn.cursor()
        try:
            c.execute("INSERT INTO results VALUES (?, ?, ?, ?, ?, ?)", (name, english, math, science, cre, sst))
            self.conn.commit()
            messagebox.showinfo("Success", "Result added successfully!")

            # Update Excel file
            self.update_excel()

            # Clear form
            self.clear_form()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Result for this student already exists. Use Update Result instead.")

    def update_excel(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM results")
        rows = c.fetchall()

        wb = load_workbook('exam_results.xlsx')
        ws = wb.active

        # Clear existing data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # Rewrite data and calculate total and average marks
        for idx, row in enumerate(rows, start=2):
            total_marks = sum(row[1:])
            average_marks = total_marks / 5
            row_data = [row[0], *row[1:], total_marks, average_marks]
            ws.append(row_data)

        # Sort by total marks in descending order
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_sort_condition("H2:H" + str(len(rows) + 1), descending=True)

        wb.save('exam_results.xlsx')

    def clear_form(self):
        self.name_entry.delete(0, tk.END)
        self.english_entry.delete(0, tk.END)
        self.math_entry.delete(0, tk.END)
        self.science_entry.delete(0, tk.END)
        self.cre_entry.delete(0, tk.END)
        self.sst_entry.delete(0, tk.END)

    def validate_marks(self, marks):
        try:
            marks = int(marks)
            if 0 <= marks <= 100:
                return marks
            else:
                return None
        except ValueError:
            return None

    def generate_report(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM results")
        rows = c.fetchall()

        report_text = ""
        for idx, row in enumerate(rows, start=1):
            total_marks = sum(row[1:])
            average_marks = total_marks / 5
            report_text += f"Position {idx}: {row[0]}\nTotal Marks: {total_marks}\nAverage Marks: {average_marks:.2f}\n\n"

        self.report_text.delete(1.0, tk.END)
        self.report_text.insert(tk.END, report_text)

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = ExamResultsSystem()
    app.run()
